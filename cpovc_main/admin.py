from django.contrib import admin
from django.http import HttpResponse
from .models import SetupGeography


def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=list_geo.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Name"),
        smart_str(u"Parent"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.area_name),
            smart_str(obj.parent_area_id),
        ])
    return response
export_csv.short_description = u"Export CSV"


def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=list_geo.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("List Geo")
    row_num = 0
    columns = [
        (u"ID", 2000),
        (u"Name", 6000),
        (u"Parent", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.area_name,
            obj.parent_area_id,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response
export_xls.short_description = u"Export XLS"


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    fmt = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=fmt)
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "List Geo"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Name", 70),
        (u"Parent", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        col_width = columns[col_num][1]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = col_width

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.area_name,
            obj.parent_area_id,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"


class GeoModelAdmin(admin.ModelAdmin):
    search_fields = ['area_id', 'area_name']
    list_display = ['area_id', 'area_name', 'area_type_id', 'area_code',
                    'parent_area_id']
    readonly_fields = ['area_id']
    list_filter = ['area_type_id']
    actions = [export_csv, export_xls, export_xlsx]

admin.site.register(SetupGeography, GeoModelAdmin)

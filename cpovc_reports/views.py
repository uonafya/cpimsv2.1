"""Views for all reports."""
import os
import re
import csv
import time
import uuid
import base64
import urllib
import string
import mimetypes
import calendar
from datetime import datetime
# from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib import messages
from django.http import JsonResponse
from .forms import CaseLoad
from .functions import (
    get_case_details, case_load_header, get_data_element,
    simple_document, draw_page, get_geo_locations, get_data, get_period,
    get_sub_county_info, get_raw_data, create_year_list, get_totals,
    get_case_data, org_unit_tree, get_performance, get_performance_detail,
    get_pivot_data, get_pivot_ovc)

from cpovc_registry.models import RegOrgUnit
from cpovc_registry.functions import get_contacts, merge_two_dicts

from cpovc_auth.models import AppUser

from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas

from django.conf import settings
from django.contrib.auth.decorators import login_required

MEDIA_ROOT = settings.MEDIA_ROOT
DOC_ROOT = settings.DOCUMENT_ROOT


@login_required
def reports_cpims(request, id):
    """Method for all other reports."""
    try:
        doc_id = int(id)
        report_name = 'CPIMS'
        form = CaseLoad(request.user, data={'report_id': doc_id})
        docs = {1: 'KNBS',
                2: 'NCCS',
                3: 'SI and CCI Population Returns',
                4: 'Health',
                5: 'Ad Hoc',
                6: 'OVC'}
        if doc_id in docs:
            report_name = docs[doc_id]
        cal_years = create_year_list(year_type='C')
        cyear_list = [a for (a, b) in cal_years]
        fyear_list = ['%s/%s' % (a, a + 1) for (a, b) in cal_years]
        return render(request, 'reports/reports_index.html',
                      {'form': form, 'status': 200, 'doc_id': doc_id,
                       'report_name': report_name, 'cyears': cyear_list,
                       'fyears': fyear_list})
    except Exception, e:
        raise e


def arrange_pending(mylist):
    """Function to push pending values at end of interventions."""
    my_len = len(mylist)
    val = None
    if my_len > 1:
        for idx, ttt in enumerate(mylist):
            if ttt[2] == 'Pending':
                mylist.remove(ttt)
                # del mylist[idx]
                val = ttt
    vals = mylist
    if val:
        vals.insert(0, val)
    return vals


@login_required
def reports_home(request):
    """Some default page for reports home page."""
    try:
        blank_date = '..' * 23
        blank_time = '.' * 20
        address = 'P.O Box %s' % ('.' * 30)
        params, location = {}, '.' * 20
        form = CaseLoad(request.user)
        if request.method == 'POST':
            doc_id = request.POST.get('document_type')
            doc_name = request.POST.get('doc_name')
            org_unit_id = request.POST.get('org_unit')
            case_id = request.POST.get('case_id')
            child_name = request.POST.get('child')
            child_id = int(request.POST.get('cpims_child'))
            # Organisation units details
            orgs = RegOrgUnit.objects.select_related().get(
                id=org_unit_id, is_void=False)
            org_contacts = get_contacts(org_unit_id)
            params['org_unit'] = orgs.org_unit_name
            params['case_id'] = case_id
            if 'contact_CPOA' in org_contacts:
                address = org_contacts['contact_CPOA'].replace('\r\n', '<br/>')
            params['address'] = str(address)
            # Physical location
            if 'contact_CPHA' in org_contacts:
                location = org_contacts['contact_CPHA'].replace(
                    '\r\n', '<br/>')
            params['physical_location'] = str(location)
            # Get geo details
            geos = get_geo_locations(org_unit_id)
            sub_county = ', '.join(geos)
            params['sub_county'] = sub_county
            params['child_id'] = child_id
            params['child_name'] = child_name.upper()
            user_id = request.user.id
            print "Params", child_name, params
            case_data = get_case_data(params)
            if case_data:
                params['case_serial'] = case_data['case_serial']
                params['case_geo'] = case_data['case_geo']
            # Other params for documents
            today = datetime.now()
            year = today.strftime('%Y')
            month = today.strftime('%b')
            params['summon_date'] = blank_date
            params['summon_time'] = blank_time
            file_name = '%s_%s_%s_%s_%s' % (
                doc_id, month, year, child_id, user_id)
            simple_document(document_name=doc_name, report_name=file_name,
                            params=params)
            results = {'file_name': file_name}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        return render(request, 'reports/reports_documents.html',
                      {'form': form, 'status': 200})
    except Exception, e:
        print 'Error writing report - %s' % (str(e))
        raise e


def write_html(data, file_name, report_variables):
    """Method to write html given data."""
    try:
        report_type = report_variables['report_type']
        row_cnt = 0
        table_string = ""
        for val in data:
            row = [str(t) for t in val]
            row_cnt += 1
            table_string += "<tr>" + \
                "<td>" + \
                string.join(row, "</td><td>") + \
                "</td>" + \
                "</tr>\n"
        html = '<table class="table table-bordered">'
        html += "<thead>"
        category_title = case_load_header(
            report_type=report_type, header=True, params=report_variables)
        category_title = category_title.format(**report_variables)
        html += "%s</thead>" % (category_title)
        html += '<tbody>%s' % (table_string)
        html += '</tbody></table>'
        return html
    except Exception, e:
        raise e


def write_xlsx(data, file_name, params):
    """Method to write xls given data."""
    try:
        report_type = params['report_region']
        xltm_tmp = params['xltm'] if 'xltm' in params else 'case_load'
        # Define some style for borders
        row_start = 9 if xltm_tmp == 'case_load' else 2
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        sheet_name = params['sheet'] if 'sheet' in params else 'Sheet'
        xls_tmp = '_orgs.xltm' if report_type == 4 else '.xltm'
        wb = load_workbook(
            '%s/%s%s' % (DOC_ROOT, xltm_tmp, xls_tmp),
            data_only=True)
        ws = wb.active if row_start == 9 else wb.get_sheet_by_name(sheet_name)
        # Lets write some data to the file
        for i, value in enumerate(data):
            for c, stats in enumerate(value):
                ws.cell(row=i + row_start, column=c + 1).value = stats
                if row_start == 9:
                    ws.cell(row=i + row_start, column=c + 1).border = border
        # Fill my placeholders with actual parameters
        if row_start == 9:
            for idx, row in enumerate(ws['A2:P5']):
                for cell in row:
                    if cell.value and "{" in cell.value:
                        print cell.value
                        cell.value = cell.value.format(**params)
                        print cell.value
        file_ext = '.xlsm' if row_start == 2 else '.xlsx'
        ws.title = sheet_name
        xls_name = '%s/%s%s' % (MEDIA_ROOT, file_name, file_ext)
        print xls_name
        wb.save(xls_name)
    except Exception, e:
        print "error writing excel - %s" % (str(e))
        raise e


def write_csv(data, file_name, params):
    """Method to write csv given data."""
    try:
        with open('%s/%s.csv' % (MEDIA_ROOT, file_name), 'wb') as csvfile:
            csvwriter = csv.writer(csvfile, delimiter=',', quotechar='"',
                                   quoting=csv.QUOTE_MINIMAL)
            csvwriter.writerows(data)
    except Exception, e:
        raise e


def write_pdf(data, file_name):
    """Method to write pdf given data."""
    try:
        pdf_name = '%s/%s.pdf' % (MEDIA_ROOT, file_name)
        doc = SimpleDocTemplate(
            pdf_name, pagesize=A4, rightMargin=30,
            leftMargin=30, topMargin=30, bottomMargin=18)
        doc.pagesize = landscape(A4)
        elements = []
        # TODO: Get this line right instead of just copying it from the
        # docs
        style = TableStyle(
            [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
             ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
             ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
             ('BACKGROUND', (0, 0), (-1, 0), colors.gray)])
        # Configure style and word wrap
        s = getSampleStyleSheet()
        s = s["BodyText"]
        s.wordWrap = 'CJK'
        data2 = [[Paragraph(cell, s) for cell in row] for row in data]
        t = Table(data2)
        t.setStyle(style)
        # Send the data and build the file
        elements.append(t)
        doc.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)

    except Exception, e:
        raise e


@login_required
def reports_caseload(request):
    """Case load views."""
    results, html = {}, None
    try:
        form = CaseLoad(request.user)
        dates = {v: k for k, v in enumerate(calendar.month_abbr)}
        if request.method == 'POST':
            sub_county_ids = request.POST.getlist('sub_county')
            sub_counties = request.POST.get('sub_county')
            county = request.POST.get('county')
            if not sub_county_ids:
                sub_county_ids = [sub_counties]
            report_type = request.POST.get('report_type')
            rperiod = request.POST.get('report_period')
            report_year = request.POST.get('report_year')
            report_region = int(request.POST.get('report_region'))
            report_unit = request.POST.get('org_unit')
            org_unit_name = request.POST.get('org_unit_name')
            results = {'res': sub_county_ids}
            case_categories = get_case_details(
                ['case_category_id', 'core_item_id', 'intervention_id',
                 'si_unit_type_id', 'cci_unit_type_id'])
            categories = {}
            for case_category in case_categories:
                case_id = case_category.item_id
                case_name = case_category.item_description
                categories[case_id] = case_name
            my_county = county if int(report_region) == 2 else False
            if report_region == 1 or report_region == 4:
                sub_county_ids = []
            sub_counties = get_sub_county_info(
                sub_county_ids, icounty=my_county)
            variables = {'sub_county_id': [], 'sub_county': []}
            for sub_county in sub_counties:
                rep_var = sub_counties[sub_county]
                variables['county'] = rep_var['county']
                variables['sub_county_id'].append(rep_var['sub_county_id'])
                variables['sub_county'].append(rep_var['sub_county'])
            # Report variables
            if report_region == 1:
                variables = {'county': 'National', 'sub_county': ['National']}
            variables['sub_county'] = ', '.join(variables['sub_county'])
            today = datetime.now()
            # year = today.strftime('%Y')
            month = today.strftime('%m')
            # Other parameters
            if report_region == 4:
                rep_unit = report_unit if report_unit else False
                variables['org_unit'] = rep_unit
                variables['county'] = ''
            else:
                variables['org_unit'] = False
            rpd = rperiod[:3] if report_type == 'M' else rperiod
            year = int(report_year) + 1 if report_type == 'M' else report_year
            if report_type == 'Q':
                report_type = rperiod.replace('tr', '')
            # Month value
            month = dates[rpd] if rpd in dates else ''
            period_params = get_period(
                report_type=report_type, year=year, month=month)
            report_variables = merge_two_dicts(variables, period_params)
            print "CASE load params ", report_variables
            # -----------------------------------------------
            ou_ids = org_unit_tree(report_unit)
            print 'OUS', len(ou_ids)
            report_variables['org_unit_tree'] = ou_ids
            all_datas = get_data(report_variables)
            all_data = all_datas['data']
            all_itvs = all_datas['itv']
            all_pending = all_datas['pending']
            report_region_name = report_variables['sub_county']
            rregion = int(report_region)
            if rregion == 2:
                report_region_name = "%s-County" % (report_variables['county'])
            elif rregion == 1:
                report_region_name = 'National'
            elif rregion == 4:
                org_uniq = org_unit_name.split()[0]
                report_region_name = 'Org-%s' % (org_uniq)
            if not all_data:
                results = {'status': 9, 'file_name': '',
                           'message': 'No data matching your query.'}
                return JsonResponse(results, content_type='application/json',
                                    safe=False)
            user_id = request.user.id
            file_name = '%s_%s_%s_0_%s' % (
                report_region_name, report_variables['label'],
                report_variables['year'], user_id)
            report_variables['org_unit_name'] = org_unit_name
            report_variables['org_units_name'] = org_unit_name
            report_variables['report_region'] = report_region
            report_variables['report_type'] = 1
            # Prepare the data
            data = get_totals(all_data, categories)
            data_itv = get_totals(all_itvs, categories)
            data_pend = get_totals(all_pending, categories)
            data.sort()
            cnt = 0
            itvs = {}
            for dsorted in data:
                cnt += 1
                dsorted.insert(0, cnt)
            for kl, dt in enumerate(data):
                dval = dt[1]
                bn = 0
                # get_interventions(data_pend, dval, bn, kl, itvs)
                for ki, itv_data in enumerate(data_itv):
                    itv_cat = itv_data[0]
                    if dval == itv_cat:
                        dts = itv_data
                        bn += 1
                        dts[0] = ''
                        dts.insert(1, '')
                    else:
                        dts = None
                    if dts:
                        # knt = '%s_%s' % (kl + 1, bn)
                        knt = kl + 1
                        if knt not in itvs:
                            itvs[knt] = [dts]
                        else:
                            itvs[knt].append(dts)
                get_interventions(data_pend, dval, bn, kl, itvs)
            fitvs = list(reversed(sorted(itvs.keys())))
            len(data)
            for idx, ttt in enumerate(fitvs):
                ffitvs = itvs[ttt]
                ffitvss = arrange_pending(ffitvs)
                for ffitv in ffitvss:
                    data.insert(ttt, ffitv)
            # Add summaries
            case_data = get_data_element()
            blank = [''] * 16
            data.append(blank)
            sum_vals = get_caseload_summary(all_datas, categories)
            for d_item in case_data:
                cval = case_data[d_item]
                d_items = sum_vals[d_item] if d_item in sum_vals else [0] * 13
                sum_val = ['', cval, ''] + d_items
                data.append(sum_val)
            # Write the data to csv
            write_csv(data, file_name, report_variables)
            # Now write html
            html = write_html(data, file_name, report_variables)
            # Lets write to pdf
            # write_pdf(data, file_name)
            # This is for xlsx
            write_xlsx(data, file_name, report_variables)
            # Output some variables for the report
            results = {'status': 0, 'file_name': file_name, 'report': html,
                       'message': 'No data matching your query.'}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        return render(request, 'reports/case_load.html',
                      {'form': form, 'results': results,
                       'report': html})
    except Exception, e:
        print 'Case load report error - %s' % (str(e))
        raise e


def get_interventions(data_itv, dval, bn, kl, itvs):
    """Method to get interventions and pending."""
    try:
        for ki, itv_data in enumerate(data_itv):
            itv_cat = itv_data[0]
            if dval == itv_cat:
                dts = itv_data
                bn += 1
                dts[0] = ''
                dts.insert(1, '')
                dts[2] = 'Pending'
            else:
                dts = None
            if dts:
                # knt = '%s_%s' % (kl + 1, bn)
                # itvs[knt] = dts
                knt = kl + 1
                if knt not in itvs:
                    itvs[knt] = [dts]
                else:
                    itvs[knt].append(dts)
        # return itvs
    except Exception, e:
        print 'Error getting intervention', str(e)
        pass


def get_caseload_summary(all_datas, categories):
    """Method to get all case load summary."""
    try:
        vals = {1: 'CHILD', 2: 'CASE', 3: 'INTV', 6: 'PEND'}
        data_keys = {1: 'data', 2: 'data', 3: 'interventions', 6: 'pending'}
        sum_vals = {}
        for val in vals:
            val_name = vals[val]
            key_name = data_keys[val]
            all_sdata = all_datas[key_name]
            if val > 1:
                val_name = 'CASE'
            if val_name in all_sdata:
                all_data = all_sdata[val_name]
                if all_data:
                    key_data = {val_name: all_data}
                    summs = get_totals(key_data, categories, val_name)
                    sum_vals[val] = summs[0]
        perc_ints = get_interven_perc(sum_vals)
        print perc_ints
        sum_vals[4] = perc_ints
        return sum_vals
    except Exception, e:
        error = 'Error getting summary - %s' % (str(e))
        print error
        return {}


def get_interven_perc(case_data):
    """Method to calculate percentage interventions."""
    try:
        int_perc = []
        cases = case_data[2]
        intervens = case_data[3]
        for val in range(0, 13):
            case = int(cases[val])
            interven = int(intervens[val])
            intp = interven * 100.0 / case if case > 0 else 0
            int_perc.append(round(intp, 1))
    except Exception, e:
        print 'Error calculating inteven percentage %s' % (str(e))
        return []
    else:
        return int_perc


def reports_generate(request):
    """Case load views."""
    results, html = {}, None
    try:
        if not request.user.is_authenticated():
            msg = 'You have been logged out. Please log in again to proceed.'
            results = {'status': 9, 'file_name': '', 'message': msg}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        dates = {v: k for k, v in enumerate(calendar.month_abbr)}
        if request.method == 'POST':
            sub_county_ids = request.POST.getlist('sub_county[]')
            sub_counties = request.POST.get('sub_county')
            county = request.POST.get('county')
            if not sub_county_ids:
                sub_county_ids = [sub_counties]
            report_type = request.POST.get('report_type')
            rperiod = request.POST.get('report_period')
            rpt_years = request.POST.get('report_year')
            rpt_iyears = request.POST.get('report_years')
            report_id = request.POST.get('report_id')
            report_unit = request.POST.get('org_unit')
            report_inst = request.POST.get('org_inst')
            org_unit_name = request.POST.get('org_unit_name')
            report_region = int(request.POST.get('report_region'))
            results = {'res': sub_county_ids}
            case_categories = get_case_details(['case_category_id'])
            institution_type = request.POST.get('institution_type')
            adhoc_id = request.POST.get('report_vars')
            org_type = request.POST.get('org_type')
            ovc_type = request.POST.get('report_ovc')
            categories = {}

            print 'ID', report_id

            if int(report_id) in [3, 4]:
                report_unit = report_inst
                rpt_years = rpt_iyears
            region_names = {1: 'National', 2: 'County',
                            3: 'Sub-county', 4: 'Organisation Unit'}
            is_fin = '/' in rpt_years
            report_year = rpt_years.split('/')[0] if is_fin else rpt_years
            for case_category in case_categories:
                case_id = case_category.item_id
                case_name = case_category.item_description
                categories[case_id] = case_name
            my_county = county if report_region == 2 else False
            if report_region == 1 or report_region == 4:
                sub_county_ids = []
            sub_counties = get_sub_county_info(
                sub_county_ids, icounty=my_county)
            variables = {'sub_county_id': [], 'sub_county': []}
            for sub_county in sub_counties:
                rep_var = sub_counties[sub_county]
                variables['county'] = rep_var['county']
                variables['sub_county_id'].append(rep_var['sub_county_id'])
                variables['sub_county'].append(rep_var['sub_county'])
            # Report variables
            if int(report_region) == 1:
                variables = {'county': 'National', 'sub_county': ['National']}
            variables['sub_county'] = ', '.join(variables['sub_county'])
            # Report variables
            variables['report_region'] = report_region
            today = datetime.now()
            # year = today.strftime('%Y')
            month = today.strftime('%m')
            # Other parameters
            if report_region == 4 or int(report_id) == 5:
                rep_unit = report_unit if report_unit else False
                variables['org_unit'] = rep_unit
            else:
                variables['org_unit'] = False
            variables['report_id'] = report_id
            rpd = rperiod[:3] if report_type == 'M' else rperiod
            ryr = rperiod[0] if report_type == 'Y' else 'F'
            if int(report_id) in [1]:
                ryr = 'C'
            year = int(report_year) + 1 if report_type == 'M' else report_year
            if report_type == 'Q':
                report_type = rperiod.replace('tr', '')
            # Month value
            month = dates[rpd] if rpd in dates else ''
            period_params = get_period(
                report_type=report_type, year=year, month=month, period=ryr)
            report_variables = merge_two_dicts(variables, period_params)
            report_variables['org_unit_name'] = org_unit_name
            report_variables['institution_type'] = institution_type
            org_type_id = org_type if org_type else 'ALL'
            report_variables['org_type'] = org_type_id
            adhoc_type = adhoc_id if adhoc_id else None
            report_variables['adhoc_type'] = adhoc_type

            report_variables['cci_si_title'] = region_names[report_region]
            sc = variables['sub_county']
            is_opt = report_region == 4
            si_c = 'Sub-County: %s' % (sc) if report_region in [2, 3] else None
            si_n = 'National' if report_region == 1 else si_c
            si_name = 'Unit: %s' % org_unit_name if is_opt else si_n
            report_variables['cci_si_name'] = si_name

            # More parameters
            inst_cats = {}
            inst_cats["TNCI"] = "Charitable Children Institution"
            inst_cats["TNSI"] = "Statutory Institution"

            case_institutions = get_case_details(
                ['si_unit_type_id', 'cci_unit_type_id'])
            for itsi in case_institutions:
                inst_id = itsi.item_id
                inst_name = itsi.item_description
                inst_cats[inst_id] = inst_name

            inst_type = institution_type if org_type_id == 'ALL' else org_type
            iname = inst_cats[inst_type] if inst_type in inst_cats else None

            icheck = iname is None and inst_type
            inst_type_name = inst_cats[inst_type] if icheck else iname
            print 'INST TYPE', inst_type_name

            unit_id = org_type_id if org_type_id != 'ALL' else institution_type
            check_region = unit_id and report_region != 4
            unit_type = 'for %s' % (inst_cats[unit_id]) if check_region else ''
            report_variables['unit_type'] = unit_type
            # all_datas = get_data(report_variables)
            # all_data = all_datas['data']
            all_data, raw_data = get_raw_data(report_variables)
            if not raw_data:
                results = {'status': 9, 'file_name': '',
                           'message': 'No data matching your query.'}
                return JsonResponse(results, content_type='application/json',
                                    safe=False)
            user_id = request.user.id
            report_cat = report_variables['sub_county']
            if report_region == 4:
                org_uniq = org_unit_name.split()[0]
                report_cat = 'Org-%s' % (org_uniq)
            elif report_region == 2:
                report_cat = "%s-County" % (report_variables['county'])
            report_cat = re.sub('[^A-Za-z0-9]+', '-', report_cat)
            file_name = '%s_%s_%s_%s_%s' % (
                report_cat, report_variables['label'],
                report_variables['year'], report_id, user_id)
            # Prepare the data
            html = all_data.format(**report_variables)
            # Write the csv
            write_csv(raw_data, file_name, report_variables)
            # Write xlsx
            if int(report_id) == 6:
                xltms = {1: 'datim', 2: 'pepfar', 3: 'datim'}
                sheets = {'datim': 'RAWDATA', 'pepfar': 'Sheet1'}
                xltm = xltms[int(ovc_type)]
                sheet = sheets[xltm]
                report_variables['xltm'] = xltm
                report_variables['sheet'] = sheet
                report_variables['sheet_start'] = 5
                del raw_data[0]
                write_xlsx(raw_data, file_name, report_variables)
            results = {'status': 0, 'file_name': file_name, 'report': html,
                       'message': 'No data matching your query.'}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        else:
            results = {'status': 0, 'file_name': None, 'report': html,
                       'message': 'Invalid request.'}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
    except Exception, e:
        print 'Error generating report - %s' % (str(e))
        raise e


@login_required
def reports_download(request, file_name):
    """Generic method for downloading files."""
    try:
        if '_' not in file_name:
            file_name = base64.urlsafe_b64decode(str(file_name))
        file_path = '%s/%s' % (MEDIA_ROOT, file_name)
        fp = open(file_path, 'rb')
        response = HttpResponse(fp.read())
        fp.close()
        mime_type, encoding = mimetypes.guess_type(file_name)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        response['Content-Type'] = mime_type
        response['Content-Length'] = str(os.stat(file_path).st_size)
        if encoding is not None:
            response['Content-Encoding'] = encoding
        # new_name = uuid.uuid3(uuid.NAMESPACE_DNS, file_name)
        # To inspect details for the below code, see
        # http://greenbytes.de/tech/tc2231/
        if u'WebKit' in request.META['HTTP_USER_AGENT']:
            # Safari 3.0 and Chrome 2.0 accepts UTF-8 encoded string
            # directly.
            filename_header = 'filename=%s' % file_name.encode(
                'utf-8')
        elif u'MSIE' in request.META['HTTP_USER_AGENT']:
            # IE does not support internationalized filename at all.
            # It can only recognize internationalized URL, so we do the
            # trick via routing rules.
            filename_header = ''
        else:
            # For others like Firefox, we follow RFC2231 (encoding
            # extension in HTTP headers).
            filename_header = 'filename*=UTF-8\'\'%s' % urllib.quote(
                file_name.encode('utf-8'))
        response['Content-Disposition'] = 'attachment; ' + filename_header
    except Exception, e:
        print "Error downloading file - %s" % (str(e))
        msg = 'Error downloading file. Please contact your administrator.'
        messages.info(request, msg)
        results, html, file_name = {}, None, None
        form = CaseLoad(request.user)
        return render(request, 'reports/reports_error.html',
                      {'form': form, 'results': results,
                       'report': html, 'file_name': file_name})
    else:
        return response


@login_required
def print_pdf(request):
    """Download without printing."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="resume.pdf"'
    p = canvas.Canvas(response)

    p.drawString(100, 100, "Some text in first page.")
    p.showPage()

    p.drawString(200, 100, "Some text in second page.")
    p.showPage()

    p.save()
    return response


@login_required
def manage_reports(request):
    """For cleaning up the reports."""
    data, users, vals = [], [], {}
    fusage = 0
    doc = "Document"
    user_id = request.user.id
    is_su = request.user.is_superuser
    try:
        if request.method == 'POST':
            report_id = request.POST.get('report_id')
            print 'Report ID ', report_id
            status, remove_msg = clean_reports(report_id)
            results = {'msg': remove_msg, 'status': status}
            return JsonResponse(results, content_type='application/json',
                                safe=False)
        rtypes = ['Case Load', 'KNBS', 'NCCS', 'Population',
                  'Health', 'Ad Hoc', 'OVC']
        doctypes = {'DSUM': 'Summon', 'DSCE': 'Social Enquiry'}
        cnt, dusage = 0, 0
        for path, dirs, files in os.walk(MEDIA_ROOT):
            for filename in files:
                full_path = os.path.join(path, filename)
                create_at = "%s" % time.ctime(os.path.getctime(full_path))
                fnames = filename.split('_')
                fsize = os.path.getsize(full_path)
                if (len(fnames) == 5):
                    cnt += 1
                    rid = int(fnames[3])
                    rep_name = fnames[0]
                    rarea = rep_name.replace('Org-U', 'OU-U')
                    report_id = uuid.uuid3(uuid.NAMESPACE_DNS, filename)
                    fname = base64.urlsafe_b64encode(filename)
                    create_vars = fnames[4].split('.')
                    create_by = int(create_vars[0])
                    ftype = (create_vars[1]).upper()
                    rperiod = '%s/%s' % (fnames[1], fnames[2])
                    # size in kb
                    rsize = "%0.1f" % (fsize / 1024.0)
                    # Templates are here also
                    if rep_name in doctypes:
                        doc_type = doctypes[rep_name]
                        rarea = "N/A"
                    else:
                        doc_type = rtypes[rid] if rid < len(rtypes) else doc
                    if create_by == user_id or is_su:
                        dusage += fsize
                        users.append(create_by)
                        data.append(
                            {'rname': fname, 'rtype': doc_type,
                             'rcreate': create_at[:-8], 'rby': create_by,
                             'rid': report_id, 'rcount': cnt,
                             'fsize': rsize, 'rarea': rarea,
                             'ftype': ftype, 'rperiod': rperiod})
        if dusage > 0:
            du_mb = dusage / (1024 * 1024.0)
            if du_mb > 1:
                fusage = "%0.1f MB" % (du_mb)
            else:
                fusage = "%0.1f KB" % (du_mb * 1024)
        if data:
            """ Query all users by this ids"""
            persons = AppUser.objects.filter(id__in=users)
            for person in persons:
                fname = person.reg_person.surname
                sname = person.reg_person.first_name
                vals[person.id] = '%s %s' % (sname, fname)
        return render(request, 'reports/reports_manage.html',
                      {'data': data, 'vals': vals, 'fusage': fusage})
    except Exception, e:
        raise e


@login_required
def manage_dashboard(request):
    """Method to manage user dashboard."""
    try:
        data = []
        dts, cts = {}, {}
        dates = []
        params = {}
        if request.method == 'POST':
            print request.POST
            user_id = request.POST.get('user_id')
            daterange = request.POST.get('daterange')
            if daterange:
                start_date, end_date = daterange.split(' - ')
                sdate_obj = datetime.strptime(start_date, '%B %d, %Y')
                edate_obj = datetime.strptime(end_date, '%B %d, %Y')
                params['start_date'] = sdate_obj
                params['end_date'] = edate_obj
                qdates = '%s to %s' % (start_date, end_date)
            else:
                end_date = datetime.now()
                start_date = end_date.replace(day=1)
                params['start_date'] = start_date
                params['end_date'] = end_date
                sdt = start_date.strftime('%B %d, %Y')
                edt = end_date.strftime('%B %d, %Y')
                qdates = '%s to %s (Current Month)' % (sdt, edt)
            children, cases = get_performance_detail(request, user_id, params)
            for case in cases:
                day = case['day']
                dates.append(day)
                dts[str(day)] = case['case_count']
            cnt = 0
            for val in children:
                date = val['created_at']
                kids = val['person_count']
                cts[str(date)] = kids
                dates.append(date)

            all_dates = list(set(dates))
            bds = sorted(all_dates)
            for bdt in bds:
                cnt += 1
                case = dts[str(bdt)] if str(bdt) in dts else 0
                kid = cts[str(bdt)] if str(bdt) in cts else 0
                dt = {'id': cnt, 'date': bdt.strftime('%a, %d-%b-%Y'),
                      'cases': case, 'children': kid}
                data.append(dt)
            print qdates

            results = {'status': 0, 'message': 'Good', 'data': data,
                       'dates': qdates}
            return JsonResponse(results, safe=False)
        persons, units, cases = get_performance(request)
        return render(request, 'reports/reports_dashboard.html',
                      {'persons': persons, 'units': units,
                       'cases': cases})
    except Exception, e:
        print 'error - %s' % (str(e))
        raise e
    else:
        pass


def clean_reports(report_id):
    """Method to clean up the reports Folder."""
    try:
        fname = base64.urlsafe_b64decode(str(report_id))
        file_name = '%s/%s' % (MEDIA_ROOT, fname)
        print file_name
        os.remove(file_name)
        return 0, 'File removed successfully.'
    except Exception, e:
        print "Delete error - %s" % (str(e))
        return 99, 'File removal error. Please contact Administrator.'


@login_required
def reports_pivot(request):
    """Method to do pivot reports."""
    try:
        form = CaseLoad(request.user)
        return render(request, 'reports/pivot.html', {'form': form})
    except Exception, e:
        raise e
    else:
        pass


@login_required
def reports_ovc_pivot(request):
    """Method to do pivot reports."""
    try:
        form = CaseLoad(request.user)
        return render(request, 'reports/pivot_ovc.html', {'form': form})
    except Exception, e:
        raise e
    else:
        pass


def reports_rawdata(request):
    """Method to do adhoc pivot reports."""
    try:
        results = get_pivot_data(request)
        return JsonResponse(results, content_type='application/json',
                            safe=False)
    except Exception, e:
        print 'error getting raw data - %s' % (str(e))
        return JsonResponse([], content_type='application/json',
                            safe=False)


def reports_ovc_rawdata(request):
    """Method to do adhoc pivot reports."""
    try:
        results = get_pivot_ovc(request)
        return JsonResponse(results, content_type='application/json',
                            safe=False)
    except Exception, e:
        print 'error getting raw data - %s' % (str(e))
        return JsonResponse([], content_type='application/json',
                            safe=False)
